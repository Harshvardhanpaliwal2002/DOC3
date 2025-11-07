# file: doc3.py
"""
CSIR Professional Research Paper Matching System
- Uses 10 template PDFs (a.pdf to j.pdf) as reference
- Matches uploaded papers based on font, style, layout, and structure
- 100% Local - No external APIs
- Professional grade accuracy with image preprocessing
"""

import os
import re
import io
import tempfile
import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
import math

import numpy as np
import cv2
import streamlit as st
from pdfminer.high_level import extract_text
from pdf2image import convert_from_path
import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Configuration ----------
TEMPLATE_PDFS = [f"{chr(97+i)}.pdf" for i in range(10)]  # a.pdf to j.pdf
CACHE_FILE = "template_features_cache.json"

# Image preprocessing parameters
GAUSSIAN_KERNEL_SIZE = 5
CLAHE_CLIP_LIMIT = 2.0
CLAHE_TILE_GRID_SIZE = (8, 8)
BILATERAL_D = 9
BILATERAL_SIGMA_COLOR = 75
BILATERAL_SIGMA_SPACE = 75

# Feature weights for matching
FONT_WEIGHT = 0.30
LAYOUT_WEIGHT = 0.25
STRUCTURE_WEIGHT = 0.25
CONTENT_WEIGHT = 0.20

# ---------- Image Preprocessing Functions ----------

def apply_gaussian_filter(image: np.ndarray) -> np.ndarray:
    """Apply Gaussian filter to reduce noise."""
    return cv2.GaussianBlur(image, (GAUSSIAN_KERNEL_SIZE, GAUSSIAN_KERNEL_SIZE), 0)

def apply_clahe(image: np.ndarray) -> np.ndarray:
    """Apply CLAHE (Contrast Limited Adaptive Histogram Equalization) to boost contrast."""
    if len(image.shape) == 3:
        lab = cv2.cvtColor(image, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP_LIMIT, tileGridSize=CLAHE_TILE_GRID_SIZE)
        l = clahe.apply(l)
        lab = cv2.merge([l, a, b])
        return cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
    else:
        clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP_LIMIT, tileGridSize=CLAHE_TILE_GRID_SIZE)
        return clahe.apply(image)

def apply_edge_detection(image: np.ndarray) -> np.ndarray:
    """Apply edge detection to highlight important features."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY) if len(image.shape) == 3 else image
    edges = cv2.Canny(gray, 50, 150)
    return edges

def apply_bilateral_filter(image: np.ndarray) -> np.ndarray:
    """Apply bilateral filter to remove noise while preserving edges."""
    if len(image.shape) == 3:
        return cv2.bilateralFilter(image, BILATERAL_D, BILATERAL_SIGMA_COLOR, BILATERAL_SIGMA_SPACE)
    else:
        return cv2.bilateralFilter(image, BILATERAL_D, BILATERAL_SIGMA_COLOR, BILATERAL_SIGMA_SPACE)

def preprocess_pdf_image(pdf_path: str, page_num: int = 0) -> np.ndarray:
    """Preprocess PDF page image with all enhancement techniques."""
    try:
        # Convert PDF page to image
        pages = convert_from_path(pdf_path, dpi=200, first_page=page_num+1, last_page=page_num+1)
        if not pages:
            return None
        
        # Convert PIL to OpenCV format
        img = np.array(pages[0])
        if len(img.shape) == 3:
            img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
        
        # Apply preprocessing pipeline
        # Step 1: Bilateral filter (preserves edges while removing noise)
        img = apply_bilateral_filter(img)
        
        # Step 2: Gaussian filter (additional noise reduction)
        img = apply_gaussian_filter(img)
        
        # Step 3: CLAHE (contrast enhancement)
        img = apply_clahe(img)
        
        return img
    except Exception as e:
        st.error(f"Image preprocessing error: {e}")
        return None

# ---------- PDF Feature Extraction Functions ----------

def extract_font_features(pdf_path: str) -> Dict:
    """Extract precise font-related features from PDF with detailed analysis."""
    features = {
        'font_families': set(),
        'font_sizes': [],
        'title_font_size': None,
        'title_font_family': None,
        'title_font_style': None,  # bold, italic, normal
        'heading_font_sizes': [],
        'heading_font_families': [],
        'heading_font_styles': [],
        'body_font_size': None,
        'body_font_family': None,
        'author_font_size': None,
        'author_font_family': None,
        'author_font_style': None,
        'abstract_font_size': None,
        'abstract_font_family': None,
        'bold_fonts': set(),
        'italic_fonts': set(),
        'font_style_distribution': defaultdict(int),
        'heading_format_details': []  # Detailed heading format info
    }
    
    try:
        doc = fitz.open(pdf_path)
        if len(doc) == 0:
            return features
        
        text_content = extract_text(pdf_path)
        text_lower = text_content.lower()
        
        # Analyze first few pages for font information
        title_found = False
        author_found = False
        abstract_found = False
        
        for page_num in range(min(5, len(doc))):  # Analyze more pages for better accuracy
            page = doc[page_num]
            blocks = page.get_text("dict")
            
            for block in blocks.get("blocks", []):
                if "lines" not in block:
                    continue
                
                for line_idx, line in enumerate(block["lines"]):
                    line_text = ""
                    line_fonts = []
                    line_sizes = []
                    line_styles = []
                    
                    for span in line.get("spans", []):
                        font_name = span.get("font", "").lower()
                        font_size = span.get("size", 0)
                        flags = span.get("flags", 0)
                        span_text = span.get("text", "").strip()
                        line_text += span_text + " "
                        
                        if font_name and font_size > 0:
                            features['font_families'].add(font_name)
                            features['font_sizes'].append(font_size)
                            line_fonts.append(font_name)
                            line_sizes.append(font_size)
                            
                            # Determine style
                            style = "normal"
                            if flags & 16:  # Bold
                                style = "bold"
                                features['bold_fonts'].add(font_name)
                            if flags & 1:  # Italic
                                style = "italic" if style == "normal" else "bold-italic"
                                features['italic_fonts'].add(font_name)
                            line_styles.append(style)
                            
                            # Track font style distribution
                            style_key = f"{font_name}_{font_size:.1f}_{style}"
                            features['font_style_distribution'][style_key] += len(span_text)
                    
                    line_text = line_text.strip()
                    line_lower = line_text.lower()
                    
                    # Detect title (usually first large text, not abstract/intro)
                    if not title_found and page_num == 0 and line_idx < 5:
                        if (len(line_text) > 15 and len(line_text) < 200 and 
                            'abstract' not in line_lower and 'introduction' not in line_lower and
                            'author' not in line_lower and '@' not in line_text):
                            if line_sizes and max(line_sizes) > 10:  # Title usually larger
                                features['title_font_size'] = max(line_sizes)
                                features['title_font_family'] = line_fonts[0] if line_fonts else None
                                features['title_font_style'] = line_styles[0] if line_styles else "normal"
                                title_found = True
                    
                    # Detect author (usually after title, contains names/institutions)
                    if not author_found and page_num == 0 and line_idx < 10:
                        if (('university' in line_lower or 'institute' in line_lower or 
                             'college' in line_lower or '@' in line_text or
                             (len(line_text.split()) <= 8 and any(c.isupper() for c in line_text)))):
                            if line_sizes:
                                features['author_font_size'] = max(line_sizes)
                                features['author_font_family'] = line_fonts[0] if line_fonts else None
                                features['author_font_style'] = line_styles[0] if line_styles else "normal"
                                author_found = True
                    
                    # Detect abstract font
                    if not abstract_found and 'abstract' in line_lower:
                        abstract_found = True
                    if abstract_found and page_num == 0 and line_idx < 15:
                        if line_sizes and line_text and 'abstract' not in line_lower:
                            features['abstract_font_size'] = max(line_sizes)
                            features['abstract_font_family'] = line_fonts[0] if line_fonts else None
                    
                    # Detect heading formats (precise analysis)
                    if any(keyword in line_lower for keyword in ['abstract', 'introduction', 'methodology', 
                                                                 'methods', 'conclusion', 'results', 'discussion']):
                        if line_sizes:
                            heading_info = {
                                'text': line_text[:50],  # First 50 chars
                                'font_size': max(line_sizes),
                                'font_family': line_fonts[0] if line_fonts else None,
                                'font_style': line_styles[0] if line_styles else "normal",
                                'position_y': line.get("bbox", [0, 0, 0, 0])[1] if "bbox" in line else 0
                            }
                            features['heading_format_details'].append(heading_info)
                            if heading_info['font_size'] not in features['heading_font_sizes']:
                                features['heading_font_sizes'].append(heading_info['font_size'])
                            if heading_info['font_family'] and heading_info['font_family'] not in features['heading_font_families']:
                                features['heading_font_families'].append(heading_info['font_family'])
                            if heading_info['font_style'] not in features['heading_font_styles']:
                                features['heading_font_styles'].append(heading_info['font_style'])
        
        # Calculate statistics
        if features['font_sizes']:
            font_sizes = sorted(set(features['font_sizes']), reverse=True)
            # Title is usually the largest font
            if not features['title_font_size']:
                features['title_font_size'] = font_sizes[0] if font_sizes else None
            # Headings are usually 2nd-4th largest
            if not features['heading_font_sizes']:
                features['heading_font_sizes'] = font_sizes[1:4] if len(font_sizes) > 1 else []
            # Body is usually the most common size
            from collections import Counter
            size_counts = Counter(features['font_sizes'])
            if not features['body_font_size']:
                features['body_font_size'] = size_counts.most_common(1)[0][0] if size_counts else None
            if not features['body_font_family']:
                # Find most common font family
                all_fonts = [f for f in features['font_families']]
                if all_fonts:
                    font_counts = Counter(all_fonts)
                    features['body_font_family'] = font_counts.most_common(1)[0][0] if font_counts else None
        
        doc.close()
        
        # Convert sets to lists for JSON serialization
        features['font_families'] = list(features['font_families'])
        features['bold_fonts'] = list(features['bold_fonts'])
        features['italic_fonts'] = list(features['italic_fonts'])
        features['font_style_distribution'] = dict(features['font_style_distribution'])
        features['heading_font_sizes'] = sorted(set(features['heading_font_sizes']), reverse=True)
        features['heading_font_families'] = list(set(features['heading_font_families']))
        features['heading_font_styles'] = list(set(features['heading_font_styles']))
        
    except Exception as e:
        st.warning(f"Font extraction error for {pdf_path}: {e}")
    
    return features

def extract_layout_features(pdf_path: str) -> Dict:
    """Extract precise layout-related features including spacing details."""
    features = {
        'page_width': None,
        'page_height': None,
        'margins': {},
        'column_count': 1,
        'line_spacing_avg': None,
        'paragraph_spacing_avg': None,
        'title_spacing_above': None,
        'title_spacing_below': None,
        'author_spacing_above': None,
        'author_spacing_below': None,
        'heading_spacing_above': None,
        'heading_spacing_below': None,
        'abstract_spacing_above': None,
        'abstract_indentation': None,
        'section_spacing': []
    }
    
    try:
        doc = fitz.open(pdf_path)
        if len(doc) == 0:
            return features
        
        page = doc[0]
        rect = page.rect
        
        features['page_width'] = rect.width
        features['page_height'] = rect.height
        
        # Estimate margins and spacing from text blocks
        blocks = page.get_text("dict")
        if blocks.get("blocks"):
            x_coords = []
            y_coords = []
            block_positions = []  # Store block positions for spacing analysis
            
            for block in blocks["blocks"]:
                if "bbox" in block:
                    bbox = block["bbox"]
                    x_coords.extend([bbox[0], bbox[2]])
                    y_coords.extend([bbox[1], bbox[3]])
                    block_positions.append({
                        'y_top': bbox[1],
                        'y_bottom': bbox[3],
                        'x_left': bbox[0],
                        'x_right': bbox[2]
                    })
            
            if x_coords and y_coords:
                features['margins'] = {
                    'left': min(x_coords) if x_coords else 0,
                    'right': rect.width - max(x_coords) if x_coords else 0,
                    'top': min(y_coords) if y_coords else 0,
                    'bottom': rect.height - max(y_coords) if y_coords else 0
                }
            
            # Calculate spacing between blocks (line spacing, paragraph spacing)
            if len(block_positions) > 1:
                spacings = []
                for i in range(len(block_positions) - 1):
                    spacing = block_positions[i+1]['y_top'] - block_positions[i]['y_bottom']
                    if spacing > 0:
                        spacings.append(spacing)
                
                if spacings:
                    features['paragraph_spacing_avg'] = sum(spacings) / len(spacings)
                    
                    # Small spacings are likely line spacing
                    small_spacings = [s for s in spacings if s < 20]
                    if small_spacings:
                        features['line_spacing_avg'] = sum(small_spacings) / len(small_spacings)
            
            # Analyze spacing around title, author, headings
            text_content = extract_text(pdf_path)
            text_lower = text_content.lower()
            
            # Find title and author positions for spacing analysis
            title_y = None
            author_y = None
            abstract_y = None
            
            for block in blocks.get("blocks", []):
                if "lines" in block:
                    for line in block["lines"]:
                        line_text = " ".join([span.get("text", "") for span in line.get("spans", [])])
                        line_lower = line_text.lower().strip()
                        
                        if not title_y and len(line_text) > 15 and len(line_text) < 200:
                            if 'abstract' not in line_lower and 'introduction' not in line_lower:
                                if "bbox" in line:
                                    title_y = line["bbox"][1]
                        
                        if not author_y and ('university' in line_lower or 'institute' in line_lower or '@' in line_text):
                            if "bbox" in line:
                                author_y = line["bbox"][1]
                        
                        if not abstract_y and 'abstract' in line_lower:
                            if "bbox" in line:
                                abstract_y = line["bbox"][1]
            
            # Calculate spacing around title
            if title_y:
                for pos in block_positions:
                    if abs(pos['y_top'] - title_y) < 50:  # Near title
                        if pos['y_top'] < title_y:
                            features['title_spacing_above'] = title_y - pos['y_bottom']
                        elif pos['y_bottom'] > title_y:
                            features['title_spacing_below'] = pos['y_top'] - title_y
                            break
            
            # Calculate spacing around author
            if author_y:
                for pos in block_positions:
                    if abs(pos['y_top'] - author_y) < 50:  # Near author
                        if pos['y_top'] < author_y:
                            features['author_spacing_above'] = author_y - pos['y_bottom']
                        elif pos['y_bottom'] > author_y:
                            features['author_spacing_below'] = pos['y_top'] - author_y
                            break
            
            # Calculate abstract indentation
            if abstract_y:
                for pos in block_positions:
                    if abs(pos['y_top'] - abstract_y) < 50:
                        features['abstract_indentation'] = pos['x_left'] - features['margins'].get('left', 0)
                        features['abstract_spacing_above'] = abstract_y - (pos.get('y_bottom', abstract_y) if pos.get('y_bottom', 0) < abstract_y else 0)
                        break
        
        doc.close()
        
    except Exception as e:
        st.warning(f"Layout extraction error for {pdf_path}: {e}")
    
    return features

def extract_structure_features(pdf_path: str) -> Dict:
    """Extract precise structural features including heading formats and spacing."""
    features = {
        'has_abstract': False,
        'has_introduction': False,
        'has_methodology': False,
        'has_conclusion': False,
        'has_references': False,
        'section_headings': [],
        'heading_formats': [],  # Detailed format of each heading
        'uses_roman_numerals': False,
        'uses_numbered_sections': False,
        'title_position': 'top',
        'author_position': 'below_title',
        'author_format': None,  # How authors are formatted
        'abstract_format': None,  # How abstract is formatted
        'heading_alignment': [],  # Left, center, right
        'heading_capitalization': []  # ALL CAPS, Title Case, Sentence case
    }
    
    try:
        text = extract_text(pdf_path)
        text_lower = text.lower()
        lines = text.split('\n')
        
        # Check for common sections with precise format detection
        section_keywords = {
            'abstract': ['abstract', 'summary'],
            'introduction': ['introduction', '1.', 'i.'],
            'methodology': ['methodology', 'methods', 'materials and methods'],
            'conclusion': ['conclusion', 'conclusions', 'discussion'],
            'references': ['references', 'bibliography', 'works cited']
        }
        
        for section, keywords in section_keywords.items():
            for keyword in keywords:
                if keyword in text_lower:
                    features[f'has_{section}'] = True
                    # Find the exact format
                    for i, line in enumerate(lines):
                        if keyword in line.lower():
                            line_stripped = line.strip()
                            # Detect capitalization style
                            if line_stripped.isupper():
                                cap_style = "ALL_CAPS"
                            elif line_stripped.istitle():
                                cap_style = "Title_Case"
                            elif line_stripped and line_stripped[0].isupper():
                                cap_style = "Sentence_case"
                            else:
                                cap_style = "lowercase"
                            
                            # Detect if numbered or has special formatting
                            has_number = bool(re.search(r'^\s*\d+\.', line_stripped))
                            has_roman = bool(re.search(r'^\s*[IVX]+\.', line_stripped))
                            
                            heading_format = {
                                'section': section,
                                'text': line_stripped[:100],
                                'capitalization': cap_style,
                                'has_number': has_number,
                                'has_roman': has_roman,
                                'length': len(line_stripped)
                            }
                            features['heading_formats'].append(heading_format)
                            features['heading_capitalization'].append(cap_style)
                            break
                    break
        
        # Check for Roman numerals with precise detection
        roman_pattern = r'\b([IVX]+)\.\s+(?:introduction|methodology|conclusion|literature|abstract|results|discussion)'
        if re.search(roman_pattern, text, re.IGNORECASE):
            features['uses_roman_numerals'] = True
        
        # Check for numbered sections
        numbered_pattern = r'\b(\d+)\.\s+(?:introduction|methodology|conclusion|literature|abstract|results|discussion)'
        if re.search(numbered_pattern, text, re.IGNORECASE):
            features['uses_numbered_sections'] = True
        
        # Extract section headings with format details
        heading_patterns = [
            r'(?:^|\n)\s*([IVX]+\.?\s+[A-Z][^\n]{5,50})\s*\n',
            r'(?:^|\n)\s*(\d+\.?\s+[A-Z][^\n]{5,50})\s*\n',
            r'(?:^|\n)\s*([A-Z][A-Z\s]{10,50})\s*\n'
        ]
        
        for pattern in heading_patterns:
            matches = re.findall(pattern, text)
            for match in matches[:10]:  # Limit to first 10
                features['section_headings'].append(match.strip())
        
        # Analyze author format
        for i, line in enumerate(lines[:15]):
            line_lower = line.lower()
            if ('university' in line_lower or 'institute' in line_lower or 
                '@' in line or 'author' in line_lower):
                line_stripped = line.strip()
                if line_stripped:
                    # Detect author format style
                    if ',' in line_stripped:
                        author_format = "comma_separated"
                    elif 'and' in line_lower:
                        author_format = "and_separated"
                    elif '@' in line_stripped:
                        author_format = "with_email"
                    else:
                        author_format = "single_line"
                    
                    features['author_format'] = author_format
                    features['author_position'] = 'below_title' if i > 2 else 'same_line'
                    break
        
        # Analyze abstract format
        abstract_started = False
        for i, line in enumerate(lines):
            if 'abstract' in line.lower():
                abstract_started = True
                # Check if abstract has indentation, special formatting
                next_lines = lines[i+1:i+4] if i+1 < len(lines) else []
                if next_lines:
                    first_para = next_lines[0].strip()
                    if first_para:
                        # Check indentation (starts with spaces)
                        if first_para.startswith('  ') or first_para.startswith('\t'):
                            features['abstract_format'] = "indented"
                        elif first_para[0].isupper():
                            features['abstract_format'] = "paragraph"
                        else:
                            features['abstract_format'] = "normal"
                break
        
        # Determine title position
        if lines:
            features['title_position'] = 'top' if len(lines) > 0 else 'unknown'
        
    except Exception as e:
        st.warning(f"Structure extraction error for {pdf_path}: {e}")
    
    return features

def find_title_and_authors_improved(text: str) -> Dict:
    """
    Improved title and author detection using scoring system.
    Based on doc2.py logic with enhancements.
    """
    # Split text into lines and analyze each line
    lines = text.split('\n')
    title_candidates = []
    
    # Look for title patterns in the first 20 lines
    for i, line in enumerate(lines[:20]):
        line = line.strip()
        if not line or len(line) < 10:  # Skip empty or very short lines
            continue
            
        # Skip lines that are clearly not titles
        if any(skip_word in line.lower() for skip_word in [
            'abstract', 'introduction', 'keywords', 'author', 'email', 
            'university', 'college', 'department', 'institute', 'correspondence',
            'received', 'accepted', 'published', 'doi:', 'http', 'www',
            'business school', 'school', 'faculty', 'professor', 'dr.', 'prof.',
            'phd', 'mba', 'msc', 'bsc', 'ma', 'ba', 'corresponding author',
            'affiliation', 'address', 'phone', 'fax', 'email:', 'e-mail',
            'intern', 'center', 'in-charge', 'nielit', 'wbl'  # Added common author/affiliation terms
        ]):
            continue
        
        # Calculate title score based on multiple factors
        score = 0
        
        # Factor 1: Length (titles are usually 20-200 characters)
        if 20 <= len(line) <= 200:
            score += 2
        elif 10 <= len(line) <= 300:
            score += 1
            
        # Factor 2: ALL CAPS (very common for academic titles)
        if line.isupper() and len(line) > 15:
            score += 3
            
        # Factor 3: Title case (First Letter Of Each Word Capitalized)
        if line.istitle() and len(line) > 20:
            score += 2
            
        # Factor 3.5: Mixed case (some words capitalized, some not - common in titles)
        words = line.split()
        if len(words) > 2:
            capitalized_words = sum(1 for word in words if word[0].isupper() and len(word) > 1)
            if capitalized_words >= len(words) * 0.5:  # At least 50% of words capitalized
                score += 1
            
        # Factor 4: Contains academic keywords
        academic_words = [
            'study', 'analysis', 'approach', 'method', 'model', 'system',
            'framework', 'algorithm', 'technique', 'evaluation', 'assessment',
            'investigation', 'research', 'development', 'implementation',
            'application', 'comparison', 'review', 'survey', 'using', 'based',
            'towards', 'for', 'of', 'in', 'on', 'and', 'the'
        ]
        if any(word in line.lower() for word in academic_words):
            score += 1
            
        # Factor 5: Position (titles are usually early in the document)
        if i < 5:
            score += 2
        elif i < 10:
            score += 1
            
        # Factor 6: Not too many numbers (titles rarely have many numbers)
        number_count = sum(1 for c in line if c.isdigit())
        if number_count < len(line) * 0.1:  # Less than 10% numbers
            score += 1
            
        # Factor 7: Not too many special characters
        special_count = sum(1 for c in line if not c.isalnum() and not c.isspace())
        if special_count < len(line) * 0.2:  # Less than 20% special chars
            score += 1
            
        # Factor 8: Line is not too long (titles are usually concise)
        if len(line) < 100:
            score += 1
        
        # Penalty factors - reduce score for lines that look like author info
        penalty = 0
        
        # Penalty 1: Contains numbers at the beginning (like "1 Business School")
        if re.match(r'^\d+\s+', line):
            penalty += 3
            
        # Penalty 2: Very short lines (likely author names)
        if len(line) < 30:
            penalty += 1
            
        # Penalty 3: Contains location names or institutional terms
        location_words = [
            'china', 'usa', 'uk', 'germany', 'france', 'japan', 'india',
            'beijing', 'shanghai', 'london', 'new york', 'tokyo', 'berlin',
            'kokrajhar', 'nielit', 'intern', 'in-charge', 'center', 'ec'
        ]
        if any(loc in line.lower() for loc in location_words):
            penalty += 3
            
        # Penalty 4: Contains academic titles/degrees
        degree_words = ['professor', 'dr.', 'prof.', 'phd', 'mba', 'msc', 'bsc']
        if any(degree in line.lower() for degree in degree_words):
            penalty += 2
        
        # Penalty 5: Contains institutional words
        institutional_words = ['school', 'university', 'college', 'institute', 'faculty', 'center', 'centre']
        if any(inst in line.lower() for inst in institutional_words):
            penalty += 3
        
        # Apply penalty
        score -= penalty
        
        # Only consider lines with decent scores
        if score >= 4:
            title_candidates.append((line, score, i))
    
    # Sort by score (highest first) and position (earlier first)
    title_candidates.sort(key=lambda x: (-x[1], x[2]))
    
    # Get the best title candidate
    title = None
    title_idx = None
    if title_candidates:
        best_title = title_candidates[0][0]
        
        # Final validation - make sure it's not author information
        if not (re.match(r'^\d+\s+', best_title) or  # Starts with number
            any(author_word in best_title.lower() for author_word in [
                'business school', 'school', 'university', 'college', 'institute',
                'professor', 'dr.', 'prof.', 'phd', 'mba', 'msc', 'bsc',
                'intern', 'in-charge', 'center', 'nielit', 'wbl', 'ec'
            ]) or
            len(best_title) < 20 or  # Too short
            best_title.isupper() and len(best_title) < 30):  # Short all-caps
            title = best_title
            title_idx = title_candidates[0][2]
            
            # Try to extend title to multiple lines if needed
            title_lines = [best_title]
            for j in range(1, 3):  # Check next 2 lines
                if title_idx + j < len(lines):
                    next_line = lines[title_idx + j].strip()
                    if (len(next_line) > 10 and 
                        not any(skip_word in next_line.lower() for skip_word in [
                            'abstract', 'introduction', 'author', 'email', 'university',
                            'business school', 'school', 'professor', 'dr.', 'prof.',
                            'intern', 'in-charge', 'center', 'nielit'
                        ]) and
                        len(' '.join(title_lines + [next_line])) < 300):
                        title_lines.append(next_line)
                    else:
                        break
            
            title = ' '.join(title_lines)
    
    # Extract authors (look for lines after title)
    authors = None
    if title and title_idx is not None:
        title_end_idx = title_idx + (len(title.split('\n')) if '\n' in title else 1)
        author_candidates = []
        
        # Look for author information in next 15 lines after title
        for i in range(title_end_idx, min(title_end_idx + 15, len(lines))):
            line = lines[i].strip()
            if not line:
                continue
                
            # Skip if it looks like abstract or other sections
            if any(skip_word in line.lower() for skip_word in [
                'abstract', 'introduction', 'keywords', 'doi:', 'http'
            ]):
                break
                
            # Look for author patterns
            if (len(line) > 5 and 
                (any(author_word in line.lower() for author_word in [
                    'university', 'college', 'department', 'institute', 'correspondence',
                    'author', 'authors', 'affiliation'
                ]) or
                ('@' in line and '.' in line) or  # Email pattern
                (len(line.split()) <= 8 and not line.isupper() and 
                 not any(skip in line.lower() for skip in ['abstract', 'introduction', 'keywords'])))):
                author_candidates.append(line)
        
        if author_candidates:
            authors = ', '.join(author_candidates[:3])  # Take first 3 author lines
    
    return {"title": title, "authors": authors}

def extract_content_features(pdf_path: str) -> Dict:
    """Extract content-based features with improved title/author detection."""
    features = {
        'title': None,
        'authors': None,
        'abstract_length': 0,
        'introduction_length': 0,
        'methodology_length': 0,
        'conclusion_length': 0,
        'total_pages': 0,
        'word_count': 0
    }
    
    try:
        text = extract_text(pdf_path)
        text = normalize_text(text)
        features['word_count'] = len(text.split())
        
        # Try font-based title extraction first (more accurate)
        title_from_font = extract_title_using_font(pdf_path, text)
        
        # Extract title and authors using improved logic
        meta = find_title_and_authors_improved(text)
        meta_title = meta.get('title')
        
        # Combine font-based and text-based title candidates
        final_title = choose_best_title(title_from_font, meta_title)
        features['title'] = final_title or meta_title or title_from_font
        features['authors'] = meta.get('authors')
        
        # Extract sections using improved logic from doc2.py
        sections = split_sections_improved(text)
        
        # Also try content-pattern-based extraction as supplement
        pattern_sections = extract_sections_by_content_patterns(text)
        for key, value in pattern_sections.items():
            if value and len(value) > 50:
                sections[key] = value
        
        if 'abstract' in sections:
            features['abstract_length'] = len(sections['abstract'])
        if 'introduction' in sections:
            features['introduction_length'] = len(sections['introduction'])
        if 'methodology' in sections:
            features['methodology_length'] = len(sections['methodology'])
        if 'conclusion' in sections:
            features['conclusion_length'] = len(sections['conclusion'])
        
        # Get page count
        doc = fitz.open(pdf_path)
        features['total_pages'] = len(doc)
        doc.close()
        
    except Exception as e:
        st.warning(f"Content extraction error for {pdf_path}: {e}")
    
    return features

def find_roman_numeral_sections(text: str) -> Dict:
    """Find sections with Roman numeral headings like 'VI. Conclusion' (from doc2.py)."""
    sections = {}
    
    # Pattern for Roman numeral headings - more comprehensive
    roman_pattern = r'^\s*([IVX]+)\.?\s+(conclusion|conclusions|discussion|summary|implications|findings|literature\s+review|related\s+work|methodology|methods|introduction|abstract|references|bibliography|literature|review)\s*[:\-\s]*\s*$'
    roman_re = re.compile(roman_pattern, flags=re.IGNORECASE | re.MULTILINE)
    matches = list(roman_re.finditer(text))
    
    for i, m in enumerate(matches):
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_name = m.group(2).strip().lower()
        content = text[start:end].strip()
        if content and len(content) > 50:
            sections[section_name] = content
    
    return sections

def split_sections_improved(text: str) -> Dict:
    """Improved section splitting with multiple patterns and validation (enhanced from doc2.py)."""
    sections = {}
    HEADINGS = [
        "abstract", "introduction", "background", "literature review", "related work",
        "methodology", "methods", "materials and methods", "experimental", "method",
        "results", "findings", "discussion", "conclusion", "conclusions", "summary",
        "references", "bibliography", "works cited"
    ]
    
    # Pattern 1: Standard headings (more flexible - allows text after heading)
    pattern = r"^\s*(?:{})\s*[:\-\s]*(?:\n|$)".format("|".join(re.escape(h) for h in HEADINGS))
    heading_re = re.compile(pattern, flags=re.IGNORECASE | re.MULTILINE)
    matches = list(heading_re.finditer(text))
    
    # Pattern 2: Numbered headings like "1. Introduction", "2. Methodology"
    numbered_pattern = r"^\s*\d+\.?\s*(?:{})\s*[:\-\s]*(?:\n|$)".format("|".join(re.escape(h) for h in HEADINGS))
    numbered_re = re.compile(numbered_pattern, flags=re.IGNORECASE | re.MULTILINE)
    numbered_matches = list(numbered_re.finditer(text))
    
    # Combine and sort all matches
    all_matches = matches + numbered_matches
    all_matches.sort(key=lambda x: x.start())
    
    # Extract sections
    for i, m in enumerate(all_matches):
        start = m.end()
        end = all_matches[i + 1].start() if i + 1 < len(all_matches) else len(text)
        section_name = m.group(0).strip().lower()
        # Clean up section name (remove numbers, colons, etc.)
        section_name = re.sub(r'^\d+\.?\s*', '', section_name)
        section_name = re.sub(r'[:\-\s]+$', '', section_name)
        content = text[start:end].strip()
        if content and len(content) > 50:  # Only include sections with substantial content
            sections[section_name] = content
    
    # Add Roman numeral sections (they take precedence) - CRITICAL FIX
    roman_sections = find_roman_numeral_sections(text)
    for key, value in roman_sections.items():
        if value and len(value) > 50:
            sections[key] = value
    
    return sections

def extract_all_features(pdf_path: str) -> Dict:
    """Extract all features from a PDF."""
    return {
        'font': extract_font_features(pdf_path),
        'layout': extract_layout_features(pdf_path),
        'structure': extract_structure_features(pdf_path),
        'content': extract_content_features(pdf_path)
    }

def normalize_text(text: str) -> str:
    """Normalize text for better extraction."""
    text = text.replace("\r", "\n")
    text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
    return text.strip()

def is_valid_title_candidate(title: Optional[str]) -> bool:
    """Check if a string looks like a valid academic paper title."""
    if not title:
        return False

    text = " ".join(title.strip().split())
    if len(text) < 15 or len(text) > 300:
        return False

    lower = text.lower()
    skip_phrases = [
        'proceedings', 'conference on', 'international conference', 'symposium',
        'workshop on', 'journal of', 'volume', 'vol.', 'issue', 'issn', 'isbn',
        'www.', 'http', 'doi', 'copyright', 'c ', 'licensed under'
    ]
    if any(phrase in lower for phrase in skip_phrases):
        return False

    words = re.findall(r"\b[\w\-]+\b", text)
    if len(words) < 3 or len(words) > 30:
        return False

    long_word_count = sum(1 for w in words if len(w) >= 5)
    if long_word_count < 2:
        return False

    letters = [c for c in text if c.isalpha()]
    if letters:
        upper_ratio = sum(1 for c in letters if c.isupper()) / len(letters)
        if upper_ratio > 0.85:
            return False

    if text.endswith((':', ';', ',')):
        return False

    return True

def title_quality_score(title: Optional[str]) -> float:
    """Score a title candidate; higher scores mean higher confidence."""
    if not title:
        return 0.0

    text = " ".join(title.strip().split())
    if not text:
        return 0.0

    lower = text.lower()
    words = re.findall(r"\b[\w\-]+\b", text)
    if not words:
        return 0.0

    score = 0.0
    word_count = len(words)
    if 6 <= word_count <= 16:
        score += 0.55
    elif 4 <= word_count <= 22:
        score += 0.35
    elif 3 <= word_count <= 26:
        score += 0.15
    else:
        score -= 0.25

    unique_long_words = len({w.lower() for w in words if len(w) >= 6})
    score += min(unique_long_words * 0.05, 0.35)

    penalty_phrases = [
        'proceedings', 'conference', 'symposium', 'journal', 'volume', 'issue',
        'copyright', 'doi', 'issn', 'isbn', 'www.', 'http', 'c '
    ]
    if any(phrase in lower for phrase in penalty_phrases):
        score -= 0.4

    letters = [c for c in text if c.isalpha()]
    if letters:
        upper_ratio = sum(1 for c in letters if c.isupper()) / len(letters)
        lower_ratio = sum(1 for c in letters if c.islower()) / len(letters)
        if upper_ratio > 0.7:
            score -= 0.3
        if lower_ratio < 0.2:
            score -= 0.2

    if text.endswith((':', ';', ',')):
        score -= 0.2

    return score

def choose_best_title(font_title: Optional[str], text_title: Optional[str]) -> str:
    """Choose the best title between font-based and text-based candidates."""
    candidates: List[Tuple[float, str]] = []
    seen = set()

    for source_bias, candidate in [(0.1, font_title), (0.0, text_title)]:
        if not candidate:
            continue
        cleaned = " ".join(candidate.strip().split())
        if not cleaned or cleaned.lower() in seen:
            continue
        if not is_valid_title_candidate(cleaned):
            base_score = title_quality_score(cleaned)
            if base_score > 0.2 and source_bias < 0.05:
                candidates.append((base_score + source_bias, cleaned))
            continue
        score = title_quality_score(cleaned) + source_bias
        candidates.append((score, cleaned))
        seen.add(cleaned.lower())

    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    return font_title or text_title or ""

def pick_section_content_robust(sections: Dict, keys: List[str]) -> str:
    """Find the best matching section content based on provided keys (enhanced robust version)."""
    best_match = ""
    best_score = 0
    
    for k in keys:
        key_lower = k.lower()
        key_words = set(key_lower.split())
        
        for sec, content in sections.items():
            if not content.strip():
                continue
                
            sec_lower = sec.lower()
            sec_words = set(sec_lower.split())
            
            # Exact match gets highest score (return immediately)
            if key_lower == sec_lower:
                return content.strip()
            
            # Check for exact word match (e.g., "method" matches "methods")
            if key_lower in sec_lower or sec_lower in key_lower:
                score = len(key_lower) / len(sec_lower) if len(sec_lower) > 0 else 0.8
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
            
            # Word-based matching for compound terms (improved)
            if key_words.intersection(sec_words):
                # Calculate score based on word overlap
                overlap = key_words.intersection(sec_words)
                score = len(overlap) / max(len(key_words), 1)
                
                # Bonus for matching important words
                important_words = {'method', 'methodology', 'conclusion', 'abstract', 'introduction'}
                if overlap.intersection(important_words):
                    score += 0.3
                
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
            
            # Handle variations (method/methods/methodology)
            if 'method' in key_lower and 'method' in sec_lower:
                score = 0.9
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
            
            if 'conclusion' in key_lower and ('conclusion' in sec_lower or 'discussion' in sec_lower):
                score = 0.9
                if score > best_score:
                    best_score = score
                    best_match = content.strip()
    
    return best_match

def is_methodology_content(content: str) -> bool:
    """Validate if content appears to be from a methodology section (relaxed validation)."""
    if not content or len(content.strip()) < 50:  # Reduced from 100 to 50
        return False
    
    content_lower = content.lower()
    
    # Look for strong methodology indicators
    strong_method_indicators = [
        'methodology', 'methods', 'materials and methods', 'experimental', 'approach', 'procedure',
        'data collection', 'data analysis', 'experiment', 'experimental design', 'research design',
        'sampling', 'participants', 'subjects', 'procedure', 'protocol', 'technique', 'algorithm',
        'implementation', 'setup', 'configuration', 'parameters', 'variables', 'measurement',
        'instruments', 'tools', 'equipment', 'software', 'hardware', 'platform', 'dataset',
        'data set', 'collected', 'performed', 'conducted', 'applied', 'used', 'utilized'
    ]
    
    # Look for weak indicators that might be in other sections
    weak_method_indicators = [
        'conclusion', 'summary', 'results', 'findings', 'discussion', 'introduction',
        'abstract', 'literature', 'background', 'related work', 'references'
    ]
    
    # Count strong methodology indicators
    strong_count = sum(1 for indicator in strong_method_indicators if indicator in content_lower)
    
    # Count weak indicators (these should be minimal in methodology)
    weak_count = sum(1 for indicator in weak_method_indicators if indicator in content_lower)
    
    # Relaxed: Must have at least 1 strong indicator (reduced from 2) and no more than 2 weak indicators (increased from 1)
    return strong_count >= 1 and weak_count <= 2

def is_conclusion_content(content: str) -> bool:
    """Validate if content appears to be from a conclusion section (relaxed validation)."""
    if not content or len(content.strip()) < 50:  # Reduced from 100 to 50
        return False
    
    content_lower = content.lower()
    
    # Look for strong conclusion indicators
    strong_concl_indicators = [
        'conclusion', 'conclude', 'summary', 'overall', 'finally', 'in summary',
        'results show', 'findings indicate', 'implications', 'recommendations',
        'future work', 'limitations', 'contribution', 'significance', 'impact',
        'this study', 'our research', 'we found', 'we conclude', 'we suggest',
        'in conclusion', 'to conclude', 'to summarize', 'in summary', 'overall',
        'the results', 'our findings', 'the study', 'this research', 'we have shown',
        'in this paper', 'this paper', 'we present', 'we propose', 'we demonstrate'
    ]
    
    # Look for weak indicators that might be in other sections
    weak_concl_indicators = [
        'method', 'approach', 'procedure', 'experiment', 'data', 'literature',
        'introduction', 'abstract', 'methodology', 'background', 'related work'
    ]
    
    # Count strong conclusion indicators
    strong_count = sum(1 for indicator in strong_concl_indicators if indicator in content_lower)
    
    # Count weak indicators (these should be minimal in conclusion)
    weak_count = sum(1 for indicator in weak_concl_indicators if indicator in content_lower)
    
    # Relaxed: Must have at least 1 strong indicator (reduced from 2) and no more than 2 weak indicators (increased from 1)
    return strong_count >= 1 and weak_count <= 2

def extract_sections_by_content_patterns(text: str) -> Dict:
    """Extract sections based on content patterns when headings are not clear (enhanced from doc2.py)."""
    sections = {}
    
    # Abstract pattern - more flexible patterns
    abstract_patterns = [
        r'(?i)(?:abstract|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:introduction|1\.|background|overview|keywords|i\.))',
        r'(?i)abstract\s*[:\-]?\s*(.*?)(?=\n\s*(?:introduction|1\.|keywords))',
        r'(?i)abstract\s+(.*?)(?=\n\s*(?:introduction|1\.))'
    ]
    for pattern in abstract_patterns:
        abstract_match = re.search(pattern, text, re.DOTALL)
        if abstract_match:
            abstract_content = abstract_match.group(1).strip()
            if len(abstract_content) > 30:  # More lenient (reduced from 50)
                sections['abstract'] = abstract_content
                break
    
    # Introduction pattern - more flexible
    intro_patterns = [
        r'(?i)(?:introduction|1\.|i\.)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:methodology|methods|2\.|literature|related work|ii\.))',
        r'(?i)introduction\s*[:\-]?\s*(.*?)(?=\n\s*(?:methodology|methods|2\.))',
        r'(?i)1\.\s*introduction\s*[:\-]?\s*(.*?)(?=\n\s*(?:methodology|methods|2\.))'
    ]
    for pattern in intro_patterns:
        intro_match = re.search(pattern, text, re.DOTALL)
        if intro_match:
            intro_content = intro_match.group(1).strip()
            if len(intro_content) > 30:  # More lenient
                sections['introduction'] = intro_content
                break
    
    # Methodology pattern - multiple patterns for better coverage (enhanced)
    method_patterns = [
        r'(?i)(?:methodology|methods|materials and methods|experimental|approach|procedure)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:results|findings|3\.|discussion|conclusion|iv\.|v\.|vi\.))',
        r'(?i)(?:[IVX]+\.?\s+)?(?:methodology|methods)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:results|findings|discussion|conclusion))',
        r'(?i)(?:III|IV|V)\.?\s+(?:methodology|methods)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:results|findings|discussion|conclusion))',
        r'(?i)(?:3|4|5)\.?\s+(?:methodology|methods)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:results|findings|discussion|conclusion))',
        r'(?i)methodology\s*[:\-]?\s*(.*?)(?=\n\s*(?:results|findings|discussion))',
        r'(?i)methods\s*[:\-]?\s*(.*?)(?=\n\s*(?:results|findings|discussion))'
    ]
    
    for pattern in method_patterns:
        method_match = re.search(pattern, text, re.DOTALL)
        if method_match:
            method_content = method_match.group(1).strip()
            if len(method_content) > 50:  # More lenient (reduced from 100)
                sections['methodology'] = method_content
                break
    
    # Conclusion pattern - more precise patterns including Roman numerals
    concl_patterns = [
        r'(?i)(?:[IVX]+\.?\s+)?(?:conclusion|conclusions|discussion|summary|implications|findings)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
        r'(?i)(?:\d+\.?\s*)?(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
        r'(?i)(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|$))',
        # Specific Roman numeral patterns
        r'(?i)(?:VI|VII|VIII|IX|X|XI|XII)\.?\s+(?:conclusion|conclusions|discussion|summary)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))'
    ]
    
    for pattern in concl_patterns:
        concl_match = re.search(pattern, text, re.DOTALL)
        if concl_match:
            content = concl_match.group(1).strip()
            if len(content) > 100 and is_conclusion_content(content):
                sections['conclusion'] = content
                break
    
    # More conservative fallback for conclusion
    if 'conclusion' not in sections:
        # Look for content near the end that has conclusion characteristics
        doc_length = len(text)
        search_start = int(doc_length * 0.85)  # Start from 85% through the document
        end_section = text[search_start:]
        
        # Look for paragraphs that look like conclusion content
        paragraphs = end_section.split('\n\n')
        concl_content = []
        for para in paragraphs:
            if len(para.strip()) > 150 and is_conclusion_content(para):
                concl_content.append(para.strip())
        if concl_content:
            sections['conclusion'] = '\n\n'.join(concl_content[:1])  # Take only the first good paragraph
    
    # Special handling for conclusion sections with Roman numerals
    if 'conclusion' not in sections:
        # Look specifically for "VI. Conclusion", "VII. Conclusion", etc.
        roman_concl_patterns = [
            r'(?i)(?:VI|VII|VIII|IX|X|XI|XII)\.?\s+(?:conclusion|conclusions)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))',
            r'(?i)(?:6|7|8|9|10|11|12)\.?\s+(?:conclusion|conclusions)\s*[:\-]?\s*\n(.*?)(?=\n\s*(?:references|bibliography|acknowledgments|$))'
        ]
        
        for pattern in roman_concl_patterns:
            concl_match = re.search(pattern, text, re.DOTALL)
            if concl_match:
                content = concl_match.group(1).strip()
                if len(content) > 100:  # More lenient for Roman numeral sections
                    sections['conclusion'] = content
                    break
    
    return sections

def extract_title_using_font(pdf_path: str, text: str) -> Optional[str]:
    """Extract title using font size analysis from PDF (improved accuracy)."""
    try:
        doc = fitz.open(pdf_path)
        if len(doc) == 0:
            return None
        
        page = doc[0]
        blocks = page.get_text("dict")
        
        title_candidates = []
        all_font_sizes = []
        
        # First pass: collect all font sizes to determine typical sizes
        for block in blocks.get("blocks", []):
            if "lines" not in block:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    font_size = span.get("size", 0)
                    if font_size > 0:
                        all_font_sizes.append(font_size)
        
        # Determine typical body font size (most common)
        if all_font_sizes:
            from collections import Counter
            size_counts = Counter(all_font_sizes)
            typical_body_size = size_counts.most_common(1)[0][0]
        else:
            typical_body_size = 12
        
        # Second pass: find title candidates (check first 20 lines for better coverage)
        for block in blocks.get("blocks", []):
            if "lines" not in block:
                continue
            
            for line_idx, line in enumerate(block.get("lines", [])):
                if line_idx >= 20:  # Check more lines (increased from 10)
                    break
                
                line_text = ""
                max_font_size = 0
                min_font_size = 999
                
                for span in line.get("spans", []):
                    span_text = span.get("text", "").strip()
                    font_size = span.get("size", 0)
                    if span_text:
                        line_text += span_text + " "
                    if font_size > 0:
                        max_font_size = max(max_font_size, font_size)
                        min_font_size = min(min_font_size, font_size)
                
                line_text = line_text.strip()
                if not line_text or len(line_text) < 10:
                    continue
                
                line_lower = line_text.lower()
                
                # Skip if it's clearly not a title
                if any(skip in line_lower for skip in [
                    'abstract', 'introduction', 'keywords', 'author', 'email',
                    'university', 'college', 'institute', 'intern', 'center',
                    'in-charge', 'nielit', 'wbl', 'received', 'accepted', 'doi',
                    'correspondence', 'affiliation', 'address'
                ]):
                    continue
                
                # Title usually has larger font size than body (relaxed: > 10pt or 20% larger than body)
                is_larger_font = max_font_size > max(10, typical_body_size * 1.2)
                
                # More lenient length requirement (15-400 characters)
                if is_larger_font and 15 <= len(line_text) <= 400:
                    # Additional validation
                    if not re.match(r'^\d+\s+', line_text):  # Doesn't start with number
                        # Calculate score based on font size difference
                        font_score = max_font_size - typical_body_size
                        title_candidates.append((line_text, font_score, max_font_size, line_idx))
        
        doc.close()
        
        # Sort by font score (largest difference first), then font size, then position
        if title_candidates:
            title_candidates.sort(key=lambda x: (-x[1], -x[2], x[3]))
            for candidate_text, _, _, _ in title_candidates:
                cleaned = " ".join(candidate_text.strip().split())
                if is_valid_title_candidate(cleaned):
                    return cleaned
            # As fallback, return highest scoring candidate if somewhat reasonable
            fallback = " ".join(title_candidates[0][0].strip().split())
            if len(fallback) >= 12:
                return fallback
        
        return None
    except Exception as e:
        return None

def extract_full_paper_data(pdf_path: str, filename: str) -> Dict:
    """Extract complete paper data for Excel export using robust extraction logic."""
    try:
        text = extract_text(pdf_path)
        text = normalize_text(text)
        
        # Try font-based title extraction first (more accurate)
        title_from_font = extract_title_using_font(pdf_path, text)
        
        # Extract title and authors using improved logic
        meta = find_title_and_authors_improved(text)
        meta_title = meta.get("title")
        
        # Combine font-based and text-based title candidates
        final_title = choose_best_title(title_from_font, meta_title)
        
        # Try heading-based extraction first
        sections = split_sections_improved(text)
        
        # Always try content-pattern-based extraction as a supplement (from doc2.py)
        pattern_sections = extract_sections_by_content_patterns(text)
        # Merge pattern-based sections with heading-based sections (pattern-based takes precedence)
        for key, value in pattern_sections.items():
            if value and len(value) > 50:  # Only add if substantial content
                sections[key] = value
        
        # Extract sections with robust matching
        abstract = pick_section_content_robust(sections, ["abstract", "summary"])
        introduction = pick_section_content_robust(sections, ["introduction", "background", "overview"])
        
        # Validate methodology content to avoid random content (but be more lenient)
        methodology = pick_section_content_robust(sections, ["methodology", "methods", "materials and methods", "experimental", "approach", "procedure", "method"])
        # Only validate if we have substantial content - be lenient
        if methodology and len(methodology.strip()) > 200:
            if not is_methodology_content(methodology):
                # Don't clear - validation might be too strict, keep it if it's substantial
                pass
        
        # Validate conclusion content (be more lenient)
        conclusion = pick_section_content_robust(sections, ["conclusion", "conclusions", "discussion", "summary", "implications", "findings"])
        if conclusion:
            # Check if it's from a Roman numeral section - be more lenient
            is_roman_section = any(roman in conclusion for roman in ['VI.', 'VII.', 'VIII.', 'IX.', 'X.', 'XI.', 'XII.'])
            if is_roman_section:
                # For Roman numeral sections, just check minimum length
                if len(conclusion.strip()) < 30:  # More lenient (reduced from 50)
                    conclusion = ""
            else:
                # For regular sections, only validate if substantial content
                if len(conclusion.strip()) > 200:
                    if not is_conclusion_content(conclusion):
                        # Don't clear - validation might be too strict, keep it if it's substantial
                        pass
                elif len(conclusion.strip()) < 30:
                    conclusion = ""
        
        references = pick_section_content_robust(sections, ["references", "bibliography", "works cited"])
        
        return {
            "file_name": filename,
            "title": final_title or meta_title or title_from_font or "",
            "authors": meta.get("authors", "") or "",
            "abstract": abstract or "",
            "introduction": introduction or "",
            "methodology": methodology or "",
            "conclusion": conclusion or "",
            "references": references or ""
        }
    except Exception as e:
        st.warning(f"Error extracting full paper data: {e}")
        import traceback
        st.error(f"Traceback: {traceback.format_exc()}")
        return {
            "file_name": filename,
            "title": "",
            "authors": "",
            "abstract": "",
            "introduction": "",
            "methodology": "",
            "conclusion": "",
            "references": ""
        }

def clean_text_for_excel(text: str) -> str:
    """Remove illegal Unicode characters for Excel compatibility."""
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    # Remove control characters that Excel doesn't like
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

def create_excel(df: pd.DataFrame) -> bytes:
    """Create a styled Excel file from extracted paper data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    header = list(df.columns)
    ws.append(header)

    for row in df.itertuples(index=False, name=None):
        safe_row = [clean_text_for_excel(str(x) if x else "") for x in row]
        ws.append(safe_row)

    # Style the header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        # Set column width (wider for content columns)
        col_name = header[col_idx - 1].lower()
        if col_name in ['abstract', 'introduction', 'methodology', 'conclusion', 'references']:
            ws.column_dimensions[get_column_letter(col_idx)].width = 60
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 40

    # Add borders to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(header)):
        for cell in row:
            cell.border = border
            # Wrap text for long content
            if cell.column_letter in [get_column_letter(i+1) for i, h in enumerate(header) if h.lower() in ['abstract', 'introduction', 'methodology', 'conclusion', 'references']]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# ---------- Template Loading and Caching ----------

@st.cache_data
def load_template_features() -> Dict[str, Dict]:
    """Load features from all template PDFs."""
    template_features = {}
    
    for template_name in TEMPLATE_PDFS:
        template_path = Path(template_name)
        if template_path.exists():
            try:
                features = extract_all_features(str(template_path))
                template_features[template_name] = features
                st.success(f"✅ Loaded template: {template_name}")
            except Exception as e:
                st.error(f"❌ Error loading {template_name}: {e}")
        else:
            st.warning(f"⚠️ Template not found: {template_name}")
    
    return template_features

# ---------- Similarity Calculation Functions ----------

def calculate_font_similarity(font1: Dict, font2: Dict) -> float:
    """Calculate precise font feature similarity (0-1) using detailed features."""
    if not font1 or not font2:
        return 0.0
    
    score = 0.0
    total_weight = 0.0
    
    # Font families similarity
    if font1.get('font_families') and font2.get('font_families'):
        set1 = set(font1['font_families'])
        set2 = set(font2['font_families'])
        if set1 or set2:
            intersection = len(set1 & set2)
            union = len(set1 | set2)
            similarity = intersection / union if union > 0 else 0.0
            score += similarity * 0.15
            total_weight += 0.15
    
    # Title font similarity (size, family, style)
    title_sim = 0.0
    if font1.get('title_font_size') and font2.get('title_font_size'):
        size1 = font1['title_font_size']
        size2 = font2['title_font_size']
        size_diff = abs(size1 - size2) / max(size1, size2) if max(size1, size2) > 0 else 1.0
        title_sim += (1.0 - min(size_diff, 1.0)) * 0.4
    
    if font1.get('title_font_family') and font2.get('title_font_family'):
        if font1['title_font_family'] == font2['title_font_family']:
            title_sim += 0.3
        else:
            title_sim += 0.1  # Partial match if similar
    
    if font1.get('title_font_style') and font2.get('title_font_style'):
        if font1['title_font_style'] == font2['title_font_style']:
            title_sim += 0.3
    
    score += title_sim * 0.20
    total_weight += 0.20
    
    # Author font similarity
    author_sim = 0.0
    if font1.get('author_font_size') and font2.get('author_font_size'):
        size1 = font1['author_font_size']
        size2 = font2['author_font_size']
        size_diff = abs(size1 - size2) / max(size1, size2) if max(size1, size2) > 0 else 1.0
        author_sim += (1.0 - min(size_diff, 1.0)) * 0.5
    
    if font1.get('author_font_family') and font2.get('author_font_family'):
        if font1['author_font_family'] == font2['author_font_family']:
            author_sim += 0.5
    
    score += author_sim * 0.15
    total_weight += 0.15
    
    # Heading font similarity (precise)
    heading_sim = 0.0
    h1_sizes = set(font1.get('heading_font_sizes', [])[:3])
    h2_sizes = set(font2.get('heading_font_sizes', [])[:3])
    if h1_sizes or h2_sizes:
        size_intersection = len(h1_sizes & h2_sizes)
        size_union = len(h1_sizes | h2_sizes)
        heading_sim += (size_intersection / size_union if size_union > 0 else 0.0) * 0.4
    
    h1_families = set(font1.get('heading_font_families', []))
    h2_families = set(font2.get('heading_font_families', []))
    if h1_families or h2_families:
        fam_intersection = len(h1_families & h2_families)
        fam_union = len(h1_families | h2_families)
        heading_sim += (fam_intersection / fam_union if fam_union > 0 else 0.0) * 0.3
    
    h1_styles = set(font1.get('heading_font_styles', []))
    h2_styles = set(font2.get('heading_font_styles', []))
    if h1_styles or h2_styles:
        style_intersection = len(h1_styles & h2_styles)
        style_union = len(h1_styles | h2_styles)
        heading_sim += (style_intersection / style_union if style_union > 0 else 0.0) * 0.3
    
    score += heading_sim * 0.15
    total_weight += 0.15
    
    # Body font size similarity
    if font1.get('body_font_size') and font2.get('body_font_size'):
        size1 = font1['body_font_size']
        size2 = font2['body_font_size']
        size_diff = abs(size1 - size2) / max(size1, size2) if max(size1, size2) > 0 else 1.0
        similarity = 1.0 - min(size_diff, 1.0)
        score += similarity * 0.15
        total_weight += 0.15
    
    # Abstract font similarity
    if font1.get('abstract_font_size') and font2.get('abstract_font_size'):
        size1 = font1['abstract_font_size']
        size2 = font2['abstract_font_size']
        size_diff = abs(size1 - size2) / max(size1, size2) if max(size1, size2) > 0 else 1.0
        similarity = 1.0 - min(size_diff, 1.0)
        score += similarity * 0.10
        total_weight += 0.10
    
    # Bold/Italic usage similarity
    bold1 = set(font1.get('bold_fonts', []))
    bold2 = set(font2.get('bold_fonts', []))
    italic1 = set(font1.get('italic_fonts', []))
    italic2 = set(font2.get('italic_fonts', []))
    
    bold_sim = len(bold1 & bold2) / len(bold1 | bold2) if (bold1 | bold2) else 0.0
    italic_sim = len(italic1 & italic2) / len(italic1 | italic2) if (italic1 | italic2) else 0.0
    
    score += (bold_sim + italic_sim) / 2 * 0.10
    total_weight += 0.10
    
    return score / total_weight if total_weight > 0 else 0.0

def calculate_layout_similarity(layout1: Dict, layout2: Dict) -> float:
    """Calculate precise layout feature similarity (0-1) including spacing details."""
    if not layout1 or not layout2:
        return 0.0
    
    score = 0.0
    total_weight = 0.0
    
    # Page dimensions similarity
    if layout1.get('page_width') and layout2.get('page_width'):
        w1, h1 = layout1['page_width'], layout1.get('page_height', 0)
        w2, h2 = layout2['page_width'], layout2.get('page_height', 0)
        
        if w1 > 0 and w2 > 0 and h1 > 0 and h2 > 0:
            w_sim = 1.0 - abs(w1 - w2) / max(w1, w2)
            h_sim = 1.0 - abs(h1 - h2) / max(h1, h2)
            score += (w_sim + h_sim) / 2 * 0.25
            total_weight += 0.25
    
    # Margins similarity
    margins1 = layout1.get('margins', {})
    margins2 = layout2.get('margins', {})
    if margins1 and margins2:
        margin_keys = ['left', 'right', 'top', 'bottom']
        margin_sims = []
        for key in margin_keys:
            if key in margins1 and key in margins2:
                m1, m2 = margins1[key], margins2[key]
                if max(m1, m2) > 0:
                    sim = 1.0 - abs(m1 - m2) / max(m1, m2)
                    margin_sims.append(sim)
        
        if margin_sims:
            score += np.mean(margin_sims) * 0.20
            total_weight += 0.20
    
    # Spacing similarity (precise)
    spacing_sims = []
    
    # Title spacing
    if layout1.get('title_spacing_below') and layout2.get('title_spacing_below'):
        s1, s2 = layout1['title_spacing_below'], layout2['title_spacing_below']
        if max(s1, s2) > 0:
            spacing_sims.append(1.0 - abs(s1 - s2) / max(s1, s2))
    
    # Author spacing
    if layout1.get('author_spacing_above') and layout2.get('author_spacing_above'):
        s1, s2 = layout1['author_spacing_above'], layout2['author_spacing_above']
        if max(s1, s2) > 0:
            spacing_sims.append(1.0 - abs(s1 - s2) / max(s1, s2))
    
    # Abstract spacing and indentation
    if layout1.get('abstract_spacing_above') and layout2.get('abstract_spacing_above'):
        s1, s2 = layout1['abstract_spacing_above'], layout2['abstract_spacing_above']
        if max(s1, s2) > 0:
            spacing_sims.append(1.0 - abs(s1 - s2) / max(s1, s2))
    
    if layout1.get('abstract_indentation') is not None and layout2.get('abstract_indentation') is not None:
        i1, i2 = layout1['abstract_indentation'] or 0, layout2['abstract_indentation'] or 0
        if max(abs(i1), abs(i2)) > 0:
            spacing_sims.append(1.0 - abs(i1 - i2) / max(abs(i1), abs(i2), 1))
        elif i1 == i2:  # Both zero or None
            spacing_sims.append(1.0)
    
    # Line and paragraph spacing
    if layout1.get('line_spacing_avg') and layout2.get('line_spacing_avg'):
        s1, s2 = layout1['line_spacing_avg'], layout2['line_spacing_avg']
        if max(s1, s2) > 0:
            spacing_sims.append(1.0 - abs(s1 - s2) / max(s1, s2))
    
    if layout1.get('paragraph_spacing_avg') and layout2.get('paragraph_spacing_avg'):
        s1, s2 = layout1['paragraph_spacing_avg'], layout2['paragraph_spacing_avg']
        if max(s1, s2) > 0:
            spacing_sims.append(1.0 - abs(s1 - s2) / max(s1, s2))
    
    if spacing_sims:
        score += np.mean(spacing_sims) * 0.30
        total_weight += 0.30
    
    # Column count similarity
    col1 = layout1.get('column_count', 1)
    col2 = layout2.get('column_count', 1)
    col_sim = 1.0 if col1 == col2 else 0.0
    score += col_sim * 0.15
    total_weight += 0.15
    
    return score / total_weight if total_weight > 0 else 0.0

def calculate_structure_similarity(struct1: Dict, struct2: Dict) -> float:
    """Calculate precise structural feature similarity (0-1) using detailed format analysis."""
    if not struct1 or not struct2:
        return 0.0
    
    score = 0.0
    total_weight = 0.0
    
    # Section presence similarity
    sections = ['abstract', 'introduction', 'methodology', 'conclusion', 'references']
    section_matches = 0
    for section in sections:
        key = f'has_{section}'
        if struct1.get(key) == struct2.get(key):
            section_matches += 1
    
    section_sim = section_matches / len(sections) if sections else 0.0
    score += section_sim * 0.25
    total_weight += 0.25
    
    # Heading format similarity (precise analysis)
    hf1 = struct1.get('heading_formats', [])
    hf2 = struct2.get('heading_formats', [])
    heading_format_sim = 0.0
    
    if hf1 and hf2:
        # Compare heading formats by section
        format_matches = 0
        total_compared = 0
        
        for h1 in hf1:
            section = h1.get('section')
            for h2 in hf2:
                if h2.get('section') == section:
                    total_compared += 1
                    # Compare capitalization
                    if h1.get('capitalization') == h2.get('capitalization'):
                        format_matches += 0.4
                    # Compare numbering style
                    if h1.get('has_number') == h2.get('has_number'):
                        format_matches += 0.3
                    if h1.get('has_roman') == h2.get('has_roman'):
                        format_matches += 0.3
                    break
        
        if total_compared > 0:
            heading_format_sim = format_matches / total_compared
    
    score += heading_format_sim * 0.25
    total_weight += 0.25
    
    # Author format similarity
    if struct1.get('author_format') and struct2.get('author_format'):
        if struct1['author_format'] == struct2['author_format']:
            score += 1.0 * 0.15
        else:
            score += 0.3 * 0.15  # Partial match
        total_weight += 0.15
    
    # Abstract format similarity
    if struct1.get('abstract_format') and struct2.get('abstract_format'):
        if struct1['abstract_format'] == struct2['abstract_format']:
            score += 1.0 * 0.10
        else:
            score += 0.3 * 0.10
        total_weight += 0.10
    
    # Roman numerals / numbered sections similarity
    roman_sim = 1.0 if struct1.get('uses_roman_numerals') == struct2.get('uses_roman_numerals') else 0.0
    numbered_sim = 1.0 if struct1.get('uses_numbered_sections') == struct2.get('uses_numbered_sections') else 0.0
    
    score += (roman_sim + numbered_sim) / 2 * 0.15
    total_weight += 0.15
    
    # Heading capitalization similarity
    cap1 = set(struct1.get('heading_capitalization', []))
    cap2 = set(struct2.get('heading_capitalization', []))
    if cap1 or cap2:
        cap_sim = len(cap1 & cap2) / len(cap1 | cap2) if (cap1 | cap2) else 0.0
        score += cap_sim * 0.10
        total_weight += 0.10
    
    return score / total_weight if total_weight > 0 else 0.0

def calculate_content_similarity(content1: Dict, content2: Dict) -> float:
    """Calculate content feature similarity (0-1)."""
    if not content1 or not content2:
        return 0.0
    
    score = 0.0
    total_weight = 0.0
    
    # Section length similarity
    sections = ['abstract', 'introduction', 'methodology', 'conclusion']
    length_sims = []
    for section in sections:
        key = f'{section}_length'
        len1 = content1.get(key, 0)
        len2 = content2.get(key, 0)
        if len1 > 0 or len2 > 0:
            max_len = max(len1, len2)
            sim = 1.0 - abs(len1 - len2) / max_len if max_len > 0 else 0.0
            length_sims.append(sim)
    
    if length_sims:
        score += np.mean(length_sims) * 0.40
        total_weight += 0.40
    
    # Page count similarity
    pages1 = content1.get('total_pages', 0)
    pages2 = content2.get('total_pages', 0)
    if pages1 > 0 and pages2 > 0:
        page_sim = 1.0 - abs(pages1 - pages2) / max(pages1, pages2)
        score += page_sim * 0.30
        total_weight += 0.30
    
    # Word count similarity
    words1 = content1.get('word_count', 0)
    words2 = content2.get('word_count', 0)
    if words1 > 0 and words2 > 0:
        word_sim = 1.0 - abs(words1 - words2) / max(words1, words2)
        score += word_sim * 0.30
        total_weight += 0.30
    
    return score / total_weight if total_weight > 0 else 0.0

def calculate_overall_similarity(features1: Dict, features2: Dict) -> Tuple[float, Dict]:
    """Calculate overall similarity score and breakdown."""
    font_sim = calculate_font_similarity(features1.get('font', {}), features2.get('font', {}))
    layout_sim = calculate_layout_similarity(features1.get('layout', {}), features2.get('layout', {}))
    struct_sim = calculate_structure_similarity(features1.get('structure', {}), features2.get('structure', {}))
    content_sim = calculate_content_similarity(features1.get('content', {}), features2.get('content', {}))
    
    overall = (
        font_sim * FONT_WEIGHT +
        layout_sim * LAYOUT_WEIGHT +
        struct_sim * STRUCTURE_WEIGHT +
        content_sim * CONTENT_WEIGHT
    )
    
    breakdown = {
        'font': font_sim * 100,
        'layout': layout_sim * 100,
        'structure': struct_sim * 100,
        'content': content_sim * 100
    }
    
    return overall, breakdown

# ---------- Matching Function ----------

def find_best_match(uploaded_features: Dict, template_features: Dict[str, Dict]) -> Tuple[str, float, Dict]:
    """Find the best matching template."""
    best_match = None
    best_score = 0.0
    best_breakdown = {}
    
    for template_name, template_feat in template_features.items():
        similarity, breakdown = calculate_overall_similarity(uploaded_features, template_feat)
        if similarity > best_score:
            best_score = similarity
            best_match = template_name
            best_breakdown = breakdown
    
    return best_match, best_score, best_breakdown

# ---------- Streamlit UI ----------

st.set_page_config(
    page_title="CSIR Research Paper Matcher",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for professional appearance
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e78;
        text-align: center;
        padding: 1rem 0;
    }
    .match-score {
        font-size: 3rem;
        font-weight: bold;
        color: #2ecc71;
        text-align: center;
    }
    .feature-box {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
        border-left: 4px solid #1f4e78;
    }
    </style>
""", unsafe_allow_html=True)

# Header with CSIR Logo
logo_path = Path("image.png")
if logo_path.exists():
    col_logo, col_text = st.columns([1, 6])
    with col_logo:
        st.image(str(logo_path), width=180)
    with col_text:
        st.markdown("""
        <div style='padding-top: 1rem; padding-left: 2px;'>
            <h1 style='font-size: 4.9rem; font-weight: bold; color: #1f4e78; margin: 0; line-height: 1.1; font-family: "Times New Roman", Times, serif;'>
                Council of Scientific and Industrial Research
            </h1>
        </div>
        """, unsafe_allow_html=True)
else:
    st.markdown('<h1 class="main-header">🔬 CSIR Professional Research Paper Matching System</h1>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.header("📋 System Information")
    st.info("""
    **Template Papers:** a.pdf to j.pdf
    
    **Matching Criteria:**
    - Font Features (30%)
    - Layout Features (25%)
    - Structure Features (25%)
    - Content Features (20%)
    """)
    
    st.header("🔄 Actions")
    if st.button("🔄 Reload Templates"):
        st.cache_data.clear()
        st.rerun()

# Main content
tab1, tab2 = st.tabs(["📤 Upload & Match", "📊 Template Information"])

with tab1:
    st.subheader("Upload Research Paper")
    uploaded_file = st.file_uploader(
        "Choose a PDF file to match against templates",
        type=['pdf'],
        help="Upload a research paper PDF to find the best matching template"
    )
    
    if uploaded_file:
        # Load template features
        with st.spinner("Loading template features..."):
            template_features = load_template_features()
        
        if not template_features:
            st.error("❌ No template PDFs found. Please ensure a.pdf to j.pdf are in the directory.")
            st.stop()
        
        st.success(f"✅ Loaded {len(template_features)} template(s)")
        
        if st.button("🔍 Analyze & Match", type="primary", use_container_width=True):
            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_file.write(uploaded_file.read())
                tmp_path = tmp_file.name
            
            try:
                # Extract features from uploaded PDF
                with st.spinner("Extracting features from uploaded paper..."):
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    status_text.info("📄 Extracting font features...")
                    progress_bar.progress(20)
                    
                    uploaded_features = extract_all_features(tmp_path)
                    
                    status_text.info("🔍 Finding best match...")
                    progress_bar.progress(80)
                    
                    # Find best match
                    best_match, match_score, breakdown = find_best_match(uploaded_features, template_features)
                    
                    progress_bar.progress(100)
                    status_text.empty()
                    progress_bar.empty()
                
                # Display results
                st.markdown("---")
                st.markdown("## 🎯 Matching Results")
                
                # Match score
                col1, col2, col3 = st.columns([1, 2, 1])
                with col2:
                    st.markdown(f'<div class="match-score">{match_score*100:.1f}%</div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="text-align: center; font-size: 1.2rem; color: #666;">Match with <strong>{best_match}</strong></div>', unsafe_allow_html=True)
                
                # Similarity breakdown
                st.markdown("### 📊 Similarity Breakdown")
                breakdown_cols = st.columns(4)
                with breakdown_cols[0]:
                    st.metric("Font Similarity", f"{breakdown['font']:.1f}%")
                with breakdown_cols[1]:
                    st.metric("Layout Similarity", f"{breakdown['layout']:.1f}%")
                with breakdown_cols[2]:
                    st.metric("Structure Similarity", f"{breakdown['structure']:.1f}%")
                with breakdown_cols[3]:
                    st.metric("Content Similarity", f"{breakdown['content']:.1f}%")
                
                # Side-by-side comparison
                st.markdown("---")
                st.markdown("### 📄 Side-by-Side Comparison")
                
                comp_col1, comp_col2 = st.columns(2)
                
                with comp_col1:
                    st.markdown(f"#### 📤 Uploaded Paper: {uploaded_file.name}")
                    # Display uploaded PDF
                    try:
                        doc = fitz.open(tmp_path)
                        if len(doc) > 0:
                            page = doc[0]
                            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                            img_data = pix.tobytes("png")
                            st.image(img_data, use_container_width=True)
                        doc.close()
                    except:
                        st.info("Preview not available")
                    
                    # Show extracted features
                    with st.expander("🔍 Extracted Features"):
                        st.json(uploaded_features)
                
                with comp_col2:
                    st.markdown(f"#### 📋 Matched Template: {best_match}")
                    # Display matched template PDF
                    try:
                        template_path = Path(best_match)
                        if template_path.exists():
                            doc = fitz.open(str(template_path))
                            if len(doc) > 0:
                                page = doc[0]
                                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                                img_data = pix.tobytes("png")
                                st.image(img_data, use_container_width=True)
                            doc.close()
                        else:
                            st.error("Template file not found")
                    except Exception as e:
                        st.error(f"Error loading template: {e}")
                    
                    # Show template features
                    with st.expander("🔍 Template Features"):
                        st.json(template_features[best_match])
                
                # All matches ranking
                st.markdown("---")
                st.markdown("### 📈 All Template Matches (Ranked)")
                
                all_matches = []
                for template_name, template_feat in template_features.items():
                    similarity, _ = calculate_overall_similarity(uploaded_features, template_feat)
                    all_matches.append({
                        'Template': template_name,
                        'Match Score (%)': f"{similarity*100:.2f}",
                        'Rank': 0
                    })
                
                all_matches.sort(key=lambda x: float(x['Match Score (%)']), reverse=True)
                for i, match in enumerate(all_matches, 1):
                    match['Rank'] = i
                
                df_matches = pd.DataFrame(all_matches)
                st.dataframe(df_matches, use_container_width=True, hide_index=True)
                
                # Extract full paper data and create Excel
                st.markdown("---")
                st.markdown("### 📊 Extracted Paper Data")
                
                with st.spinner("Extracting complete paper data..."):
                    paper_data = extract_full_paper_data(tmp_path, uploaded_file.name)
                    
                    # Create DataFrame
                    df_paper = pd.DataFrame([paper_data])
                    
                    # Display extracted data
                    st.success("✅ Paper data extracted successfully!")
                    
                    # Show preview
                    with st.expander("📋 Preview Extracted Data", expanded=True):
                        st.dataframe(df_paper, use_container_width=True)
                    
                    # Create Excel file
                    excel_bytes = create_excel(df_paper)
                    
                    # Download button
                    st.download_button(
                        label="⬇️ Download Extracted Data as Excel",
                        data=excel_bytes,
                        file_name=f"extracted_paper_data_{uploaded_file.name.replace('.pdf', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
                
            except Exception as e:
                st.error(f"❌ Error processing PDF: {e}")
                import traceback
                st.code(traceback.format_exc())
            finally:
                # Clean up temp file
                try:
                    os.unlink(tmp_path)
                except:
                    pass

with tab2:
    st.subheader("Template Papers Information")
    
    template_features = load_template_features()
    
    if template_features:
        selected_template = st.selectbox("Select Template to View", list(template_features.keys()))
        
        if selected_template:
            features = template_features[selected_template]
            
            st.markdown(f"### 📄 {selected_template}")
            
            # Display template PDF preview
            try:
                template_path = Path(selected_template)
                if template_path.exists():
                    doc = fitz.open(str(template_path))
                    if len(doc) > 0:
                        page = doc[0]
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                        img_data = pix.tobytes("png")
                        st.image(img_data, use_container_width=True, caption=f"First page of {selected_template}")
                    doc.close()
            except Exception as e:
                st.warning(f"Could not display preview: {e}")
            
            # Display features
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 🔤 Font Features")
                font_feat = features.get('font', {})
                st.json(font_feat)
            
            with col2:
                st.markdown("#### 📐 Layout Features")
                layout_feat = features.get('layout', {})
                st.json(layout_feat)
            
            col3, col4 = st.columns(2)
            
            with col3:
                st.markdown("#### 🏗️ Structure Features")
                struct_feat = features.get('structure', {})
                st.json(struct_feat)
            
            with col4:
                st.markdown("#### 📝 Content Features")
                content_feat = features.get('content', {})
                st.json(content_feat)
    else:
        st.warning("No templates loaded. Please ensure a.pdf to j.pdf are in the directory.")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #999; padding: 2rem 0;'>
    <p><strong>CSIR Research Paper Matching System</strong></p>
    <p>Professional • Accurate • Confidential</p>
</div>
""", unsafe_allow_html=True)

